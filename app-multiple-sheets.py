import pandas as pd 
import os
import sys
import datetime
import numpy as np
from PySide import QtCore, QtGui
import xlrd
from interface2 import Ui_MainWindow
import locale
from pandas import ExcelWriter
from openpyxl import Workbook, load_workbook
import openpyxl as op

def changetime(i):

    #print(type(i))
    format1 = "%H:%M:%S"
    timestr = str(i)
    #print(type(timestr))
    date = datetime.datetime.strptime(timestr, format1)
    date = date.strftime("%I:%M %p")

    date = str(date)
    
    return date


class Main_Window(QtGui.QMainWindow, Ui_MainWindow):

    def __init__(self, parent=None):

        QtGui.QMainWindow.__init__(self, parent)

        # defines all UIs based on converted *.ui file
        self.setupUi(self)

        # setup action listener
        self.setupUIAction()
        
        self.selected_path = None

    def setupUIAction(self):
        self.uploadBtn.clicked.connect(self.onSelectFolderBtnClicked)
        self.submitBtn.clicked.connect(self.onSubmitBtnClicked)
        self.submitBtn.setEnabled(False)
        

    def onSelectFolderBtnClicked(self):

        fileName = QtGui.QFileDialog.getOpenFileNames(
            self,
            "Open Excel file only",
            ".",
            "Text files (*.xlsx)"
        )[0]

        print("File Name = {}".format(fileName))

        if fileName:
            self.selected_path = fileName
            fileshow = self.selected_path
            fileshow = ''.join(fileshow)
            fileshow = os.path.basename(fileshow)

            self.file.setText("{}".format(fileshow))
            self.statuslabel.setText("READY TO CONVERT")

        self.submitBtn.setEnabled(True)

    def onSubmitBtnClicked(self):
        try:
            
            if self.selected_path:
                #Convert default excel to csv format
                
                file = self.selected_path
                file = ''.join(file)
                #print("latest file path : {}".format(file))

                #df =  pd.ExcelFile(file, engine='xlrd')
                #df = pd.read_excel('File.xlsx', sheetname='Sheet1')
                #df = xlrd.open_workbook(file)

                df = pd.read_excel(file)

                filenameonly1 = ''.join(file)
                path = os.path.basename(filenameonly1)
                folder        = os.path.dirname(filenameonly1)

                #Create New Empty Excel File
                file_name2 = "{}\Attendance.xlsx".format(folder)
                filepath = file_name2
                wb = op.Workbook()
                wb.save(filepath)

                #print("latest file with format : {}".format(path))
                #print("latest folder : {}".format(folder))

                filenameonly = path.split('.')[0]
                #print("latest file without format : {}".format(filenameonly))

                csvfile = str(".\{}.csv".format(filenameonly))

                df.to_csv(csvfile, sep=",")

                #Read CSV file
                df = pd.read_csv(csvfile) 
                #print(df.head())
                del df['Unnamed: 0']

                #Change string to datetime format
                df['Time'] = pd.to_datetime(df.Time)

                #Split datetime to date & time
                df['new_date'] = [d.date() for d in df['Time']]
                df['new_time'] = [d.time() for d in df['Time']]

                #Convert 24Hour - 12 Hour

                datelist = []
                for i in df['new_time']:
                    #print("masuk sini takkkk")
                    #print("Time before {}".format(i))
                    #format1 = "%H:%M:%S"
                    #date = datetime.datetime.strptime(str(i), format1)
                    #date = datetime.datetime.strptime(str(i), '%H:%M:%S').strftime(format1)
                    #date = date.strftime("%-I:%M %p")
                    #date = datetime.datetime.strptime(str(i), format1)
                    #date = date.strftime("%-I:%M %p")
                    #print("masuk sini before date")
                    #print(i)
                    date = changetime(i)
                    #print(date)
                    
                    datelist.append(date)

                #print("Time after {}".format(str(date)))
                df['new_time'] = pd.DataFrame({'new_time':datelist})
                #print("masuk sini lagi takkk")

                #Remove existing Time col
                #df = df.drop(columns=['Time'])

                #Get unique dates & names
                dates = np.unique(df['new_date'])
                names = np.unique(df['Name'])

                #Change all column to str
                df = df.applymap(str)

                #Define new df
                dfbaru = pd.DataFrame()

                global validco, validci

                cols = ['name', 'staffid', 'clock-in', 'clock-out', 'remarks', 'date']
                lst = []
                
                for c, date in enumerate(dates):
                    date = str(date)
                    file_name = "{}\Attendance_date_{}.csv".format(folder,str(date))
                    sheetname = "Date_{}".format(str(date))
                    #file_name2 = "{}\Attendance.xlsx".format(folder)
                    lst = []
                    for count, name in enumerate(names):
                        clockin, clockout = str('-'), str('-')
                        row = df.loc[(df['Name'] == name) & (df['new_date'] == str(date))]
                        
                        staffid = np.array(row['AC-No.'])
                        staffid = set(staffid)
                        staffid = ''.join(staffid)

                        time = np.array(row['new_time'])
                        times = time.tolist()
                        
                        validco, validci = False, False
                        #print(row)
                        for time in times :
                            
                            #print(time)
                            if "AM" in time:
                                if validci == False :
                                    clockin = time
                                    validci = True

                                    
                            elif "PM" in time:
                            
                                if validco == False :
                                    clockout = time
                                    validco = True
                                    
                            else :
                                validci = False
                                validco = False
                        #print(clockout)
                        
                            
                        if (validci == True) & (validco == True):
                            remarks = "Valid"

                        elif (validci == True) | (validco == True):
                            remarks = "Non Valid"
                            
                        elif (validci == False) & (validco == False):
                            remarks = "-"
                            clockin = str('-')
                            clockout = str('-')
                        
                        lst.append([name, staffid, clockin, clockout, remarks, str(date)])
                        

                        df1 = pd.DataFrame(lst, columns=cols)

                        

                        
                        
                        try:
                            df1.to_csv(file_name,sep = ',',index=False, header=True)
                            
                            #riter = pd.ExcelWriter(file_name2, engine='openpyxl')
                            #df1.to_excel(writer, sheetname,  index=False)
                            #writer.save()

                            book = load_workbook(file_name2)
                            writer = pd.ExcelWriter(file_name2, engine='openpyxl') 
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                            df1.to_excel(writer, sheetname,  index=False)
                            writer.save()

                            

                            self.statuslabel.setText("SUCCESS")
                            
                            
                        except:
                            self.statuslabel.setText("ERROR!")
             
            os.remove(csvfile)
            
        except BaseException as e:
            QtGui.QMessageBox.warning(None, 'Error', str(e))
            print('Error {}'.format( str(e)))
            return

if __name__ == '__main__':
    
    #datee = "08:42:00"
    #datee = changetime(datee)
    #print(type(datee))
    #print("Dalam ni Convert {}".format(datee))

    # Create Qt app
    app = QtGui.QApplication(sys.argv)
    locale.setlocale(locale.LC_ALL, 'C')

    # Create the widget and show it
    gui = Main_Window()
    gui.show()

    # Run the app
    sys.exit(app.exec_())